import io
import csv
from datetime import datetime, date, timedelta, timezone
from calendar import month_name, month_abbr
from typing import Optional, List, Dict, Any, Tuple
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import select, func, extract

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    KeepTogether,
    HRFlowable,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfgen import canvas

from app.models.income import Income
from app.models.expense import Expense
from app.models.account import Account
from app.models.user import User
from app.schemas.report import (
    ReportSummaryMetrics,
    ReportCategoryBreakdown,
    ReportSourceBreakdown,
    ReportAccountBreakdown,
    ReportTimelineItem,
    ReportTransactionItem,
    ReportDataResponse,
    ReportUserOption,
)


def parse_date_boundaries(
    period_type: str = "month",
    month: Optional[int] = None,
    year: Optional[int] = None,
    start_date_str: Optional[str] = None,
    end_date_str: Optional[str] = None,
) -> Tuple[Optional[datetime], Optional[datetime], str]:
    """
    Computes (start_datetime, end_datetime, period_label) based on provided filter parameters.
    """
    now = datetime.now(timezone.utc)
    target_year = year if year is not None and year > 0 else now.year

    if period_type == "month":
        target_month = month if month is not None and 1 <= month <= 12 else now.month
        start_dt = datetime(target_year, target_month, 1, 0, 0, 0)
        if target_month == 12:
            end_dt = datetime(target_year + 1, 1, 1, 0, 0, 0)
        else:
            end_dt = datetime(target_year, target_month + 1, 1, 0, 0, 0)
        label = f"{month_name[target_month]} {target_year}"
        return start_dt, end_dt, label

    elif period_type == "year":
        start_dt = datetime(target_year, 1, 1, 0, 0, 0)
        end_dt = datetime(target_year + 1, 1, 1, 0, 0, 0)
        label = f"Year {target_year}"
        return start_dt, end_dt, label

    elif period_type == "custom":
        if start_date_str:
            try:
                start_dt = datetime.strptime(start_date_str, "%Y-%m-%d")
            except ValueError:
                start_dt = datetime(now.year, 1, 1)
        else:
            start_dt = datetime(now.year, 1, 1)

        if end_date_str:
            try:
                # end of specified date
                parsed_end = datetime.strptime(end_date_str, "%Y-%m-%d")
                end_dt = parsed_end + timedelta(days=1)
            except ValueError:
                end_dt = now + timedelta(days=1)
        else:
            end_dt = now + timedelta(days=1)

        label = f"{start_dt.strftime('%b %d, %Y')} - {(end_dt - timedelta(seconds=1)).strftime('%b %d, %Y')}"
        return start_dt, end_dt, label

    elif period_type == "all":
        label = "All Time"
        return None, None, label

    else:
        # Fallback to current month
        target_month = now.month
        start_dt = datetime(now.year, target_month, 1, 0, 0, 0)
        end_dt = datetime(now.year + 1, 1, 1, 0, 0, 0) if target_month == 12 else datetime(now.year, target_month + 1, 1, 0, 0, 0)
        label = f"{month_name[target_month]} {now.year}"
        return start_dt, end_dt, label


def get_available_filter_options(db: Session, user_id: Optional[int] = None) -> Tuple[List[int], List[str], List[str]]:
    """Retrieve distinct years, categories, and accounts for filter dropdowns."""
    # Years from income & expense dates
    inc_years_stmt = select(extract('year', Income.date)).distinct()
    exp_years_stmt = select(extract('year', Expense.date)).distinct()
    cat_stmt = select(Expense.category).distinct()
    acct_stmt = select(Account.bank_name).distinct()
    inc_acct_stmt = select(Income.account).distinct()
    exp_acct_stmt = select(Expense.account).distinct()

    if user_id is not None:
        inc_years_stmt = inc_years_stmt.where(Income.user_id == user_id)
        exp_years_stmt = exp_years_stmt.where(Expense.user_id == user_id)
        cat_stmt = cat_stmt.where(Expense.user_id == user_id)
        acct_stmt = acct_stmt.where(Account.user_id == user_id)
        inc_acct_stmt = inc_acct_stmt.where(Income.user_id == user_id)
        exp_acct_stmt = exp_acct_stmt.where(Expense.user_id == user_id)
    
    inc_years = db.execute(inc_years_stmt).scalars().all()
    exp_years = db.execute(exp_years_stmt).scalars().all()
    
    all_years = set([int(y) for y in (list(inc_years) + list(exp_years)) if y is not None])
    all_years.add(datetime.now().year)
    sorted_years = sorted(list(all_years), reverse=True)

    # Categories from expenses
    categories = sorted([c for c in db.execute(cat_stmt).scalars().all() if c])

    # Accounts from user accounts + income/expense records
    accts_set = set([a for a in db.execute(acct_stmt).scalars().all() if a])
    for a in db.execute(inc_acct_stmt).scalars().all():
        if a:
            accts_set.add(a)
    for a in db.execute(exp_acct_stmt).scalars().all():
        if a:
            accts_set.add(a)
            
    sorted_accounts = sorted(list(accts_set)) if accts_set else ["Cash"]

    return sorted_years, categories, sorted_accounts


def get_user_report_data(
    db: Session,
    user_id: Optional[int] = None,
    period_type: str = "month",
    month: Optional[int] = None,
    year: Optional[int] = None,
    start_date_str: Optional[str] = None,
    end_date_str: Optional[str] = None,
    transaction_type: str = "all",
    category: Optional[str] = None,
    account: Optional[str] = None,
    is_limited: bool = False,
    is_admin: bool = False,
) -> ReportDataResponse:
    """
    Extracts, filters, and calculates summary metrics and full transaction history for reports.
    If user_id is None, aggregates system-wide metrics across all users (for admin audit).
    """
    start_dt, end_dt, period_label = parse_date_boundaries(
        period_type=period_type,
        month=month,
        year=year,
        start_date_str=start_date_str,
        end_date_str=end_date_str,
    )

    # Query Incomes
    inc_stmt = select(Income).options(joinedload(Income.user))
    if user_id is not None:
        inc_stmt = inc_stmt.where(Income.user_id == user_id)
    if start_dt is not None:
        inc_stmt = inc_stmt.where(Income.date >= start_dt)
    if end_dt is not None:
        inc_stmt = inc_stmt.where(Income.date < end_dt)
    if account:
        inc_stmt = inc_stmt.where(Income.account == account)
    if category:
        # Category filter on income compares with source
        inc_stmt = inc_stmt.where(Income.source == category)

    incomes: List[Income] = []
    if transaction_type in ("all", "income"):
        incomes = list(db.execute(inc_stmt.order_by(Income.date.desc())).scalars().all())

    # Query Expenses
    exp_stmt = select(Expense).options(joinedload(Expense.user))
    if user_id is not None:
        exp_stmt = exp_stmt.where(Expense.user_id == user_id)
    if start_dt is not None:
        exp_stmt = exp_stmt.where(Expense.date >= start_dt)
    if end_dt is not None:
        exp_stmt = exp_stmt.where(Expense.date < end_dt)
    if account:
        exp_stmt = exp_stmt.where(Expense.account == account)
    if category:
        exp_stmt = exp_stmt.where(Expense.category == category)

    expenses: List[Expense] = []
    if transaction_type in ("all", "expense"):
        expenses = list(db.execute(exp_stmt.order_by(Expense.date.desc())).scalars().all())

    # Calculate Totals & Metrics
    total_income = sum([float(i.amount or 0) for i in incomes])
    total_expenses = sum([float(e.amount or 0) for e in expenses])
    net_savings = total_income - total_expenses
    savings_rate = round((net_savings / total_income * 100), 1) if total_income > 0 else 0.0

    income_count = len(incomes)
    expense_count = len(expenses)
    total_transactions_count = income_count + expense_count

    avg_income_txn = round(total_income / income_count, 2) if income_count > 0 else 0.0
    avg_expense_txn = round(total_expenses / expense_count, 2) if expense_count > 0 else 0.0
    max_income = max([float(i.amount or 0) for i in incomes], default=0.0)
    max_expense = max([float(e.amount or 0) for e in expenses], default=0.0)

    # Category Breakdown (Expenses)
    cat_map: Dict[str, Dict[str, Any]] = {}
    for exp in expenses:
        cat = exp.category or "Other"
        if cat not in cat_map:
            cat_map[cat] = {"amount": 0.0, "count": 0}
        cat_map[cat]["amount"] += float(exp.amount or 0)
        cat_map[cat]["count"] += 1

    category_breakdown = []
    for cat, data in sorted(cat_map.items(), key=lambda item: item[1]["amount"], reverse=True):
        pct = round((data["amount"] / total_expenses * 100), 1) if total_expenses > 0 else 0.0
        category_breakdown.append(
            ReportCategoryBreakdown(
                category=cat,
                amount=round(data["amount"], 2),
                percentage=pct,
                count=data["count"],
            )
        )

    # Source Breakdown (Incomes)
    src_map: Dict[str, Dict[str, Any]] = {}
    for inc in incomes:
        src = inc.source or "Other"
        if src not in src_map:
            src_map[src] = {"amount": 0.0, "count": 0}
        src_map[src]["amount"] += float(inc.amount or 0)
        src_map[src]["count"] += 1

    source_breakdown = []
    for src, data in sorted(src_map.items(), key=lambda item: item[1]["amount"], reverse=True):
        pct = round((data["amount"] / total_income * 100), 1) if total_income > 0 else 0.0
        source_breakdown.append(
            ReportSourceBreakdown(
                source=src,
                amount=round(data["amount"], 2),
                percentage=pct,
                count=data["count"],
            )
        )

    # Account Breakdown
    acct_map: Dict[str, Dict[str, Any]] = {}
    for inc in incomes:
        ac = inc.account or "Cash"
        if ac not in acct_map:
            acct_map[ac] = {"income": 0.0, "expense": 0.0, "count": 0}
        acct_map[ac]["income"] += float(inc.amount or 0)
        acct_map[ac]["count"] += 1

    for exp in expenses:
        ac = exp.account or "Cash"
        if ac not in acct_map:
            acct_map[ac] = {"income": 0.0, "expense": 0.0, "count": 0}
        acct_map[ac]["expense"] += float(exp.amount or 0)
        acct_map[ac]["count"] += 1

    account_breakdown = []
    for ac, data in sorted(acct_map.items(), key=lambda item: (item[1]["income"] + item[1]["expense"]), reverse=True):
        account_breakdown.append(
            ReportAccountBreakdown(
                account=ac,
                income_amount=round(data["income"], 2),
                expense_amount=round(data["expense"], 2),
                net_amount=round(data["income"] - data["expense"], 2),
                transaction_count=data["count"],
            )
        )

    # Timeline Breakdown (Daily or Monthly series)
    is_daily = (period_type == "month") or (start_dt and end_dt and (end_dt - start_dt).days <= 45)
    time_map: Dict[str, Dict[str, Any]] = {}

    for inc in incomes:
        if inc.date:
            key = inc.date.strftime("%Y-%m-%d") if is_daily else inc.date.strftime("%Y-%m")
            label = inc.date.strftime("%d %b") if is_daily else inc.date.strftime("%b %Y")
            if key not in time_map:
                time_map[key] = {"label": label, "date_key": key, "income": 0.0, "expenses": 0.0}
            time_map[key]["income"] += float(inc.amount or 0)

    for exp in expenses:
        if exp.date:
            key = exp.date.strftime("%Y-%m-%d") if is_daily else exp.date.strftime("%Y-%m")
            label = exp.date.strftime("%d %b") if is_daily else exp.date.strftime("%b %Y")
            if key not in time_map:
                time_map[key] = {"label": label, "date_key": key, "income": 0.0, "expenses": 0.0}
            time_map[key]["expenses"] += float(exp.amount or 0)

    timeline_breakdown = []
    for key in sorted(time_map.keys()):
        item = time_map[key]
        inc_val = round(item["income"], 2)
        exp_val = round(item["expenses"], 2)
        timeline_breakdown.append(
            ReportTimelineItem(
                label=item["label"],
                date_key=item["date_key"],
                income=inc_val,
                expenses=exp_val,
                net=round(inc_val - exp_val, 2),
            )
        )

    # Build Unified Sorted Transaction List
    unified_txns: List[ReportTransactionItem] = []
    for inc in incomes:
        unified_txns.append(
            ReportTransactionItem(
                id=inc.id,
                type="income",
                date=inc.date,
                category_or_source=inc.source,
                description=f"Income from {inc.source}",
                account=inc.account or "Cash",
                amount=float(inc.amount or 0),
                user_id=inc.user_id,
                user_email=inc.user.email if inc.user else None,
            )
        )

    for exp in expenses:
        unified_txns.append(
            ReportTransactionItem(
                id=exp.id,
                type="expense",
                date=exp.date,
                category_or_source=exp.category,
                description=exp.description or f"{exp.category} Expense",
                account=exp.account or "Cash",
                amount=float(exp.amount or 0),
                user_id=exp.user_id,
                user_email=exp.user.email if exp.user else None,
            )
        )

    unified_txns.sort(key=lambda t: t.date, reverse=True)
    if is_limited:
        unified_txns = unified_txns[:10]

    # Available filter dropdown items
    available_years, available_categories, available_accounts = get_available_filter_options(db, user_id)

    available_users: Optional[List[ReportUserOption]] = None
    if is_admin:
        all_users = db.query(User).order_by(User.email.asc()).all()
        available_users = [
            ReportUserOption(
                id=u.id,
                email=u.email,
                full_name=u.profile.full_name if u.profile else None,
                role=u.role or "user",
            )
            for u in all_users
        ]

    summary = ReportSummaryMetrics(
        total_income=round(total_income, 2),
        total_expenses=round(total_expenses, 2),
        net_savings=round(net_savings, 2),
        savings_rate=savings_rate,
        total_transactions_count=total_transactions_count,
        income_count=income_count,
        expense_count=expense_count,
        avg_income_transaction=avg_income_txn,
        avg_expense_transaction=avg_expense_txn,
        max_income=round(max_income, 2),
        max_expense=round(max_expense, 2),
        period_label=period_label,
        start_date=start_dt.strftime("%Y-%m-%d") if start_dt else None,
        end_date=(end_dt - timedelta(seconds=1)).strftime("%Y-%m-%d") if end_dt else None,
    )

    return ReportDataResponse(
        summary=summary,
        category_breakdown=category_breakdown,
        source_breakdown=source_breakdown,
        account_breakdown=account_breakdown,
        timeline_breakdown=timeline_breakdown,
        transactions=unified_txns,
        available_years=available_years,
        available_categories=available_categories,
        available_accounts=available_accounts,
        available_users=available_users,
        selected_user_id=user_id,
    )


def generate_excel_report(
    db: Session,
    user: User,
    period_type: str = "month",
    month: Optional[int] = None,
    year: Optional[int] = None,
    start_date_str: Optional[str] = None,
    end_date_str: Optional[str] = None,
    transaction_type: str = "all",
    category: Optional[str] = None,
    account: Optional[str] = None,
    is_limited: bool = False,
    transactions_only: bool = False,
    target_user_id: Optional[int] = None,
) -> io.BytesIO:
    """
    Generates a professionally styled Excel workbook (.xlsx).
    If transactions_only=True, generates a standalone Transactions Ledger workbook.
    Otherwise generates the comprehensive 4-sheet financial audit workbook.
    """
    effective_user_id = target_user_id if target_user_id is not None else (user.id if user.role != "admin" else None)
    report_data = get_user_report_data(
        db=db,
        user_id=effective_user_id,
        period_type=period_type,
        month=month,
        year=year,
        start_date_str=start_date_str,
        end_date_str=end_date_str,
        transaction_type=transaction_type,
        category=category,
        account=account,
        is_limited=is_limited,
        is_admin=(user.role == "admin"),
    )

    wb = openpyxl.Workbook()

    # Style definitions
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    subheader_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    kpi_income_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    kpi_expense_fill = PatternFill(start_color="FFE4E6", end_color="FFE4E6", fill_type="solid")
    kpi_net_fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    accent_bar_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    total_row_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    font_title = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="94A3B8")
    font_section = Font(name="Segoe UI", size=12, bold=True, color="1E293B")
    font_tbl_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=10, color="1E293B")
    font_data_bold = Font(name="Segoe UI", size=10, bold=True, color="1E293B")
    font_income = Font(name="Segoe UI", size=10, bold=True, color="15803D")
    font_expense = Font(name="Segoe UI", size=10, bold=True, color="BE123C")
    font_kpi_label = Font(name="Segoe UI", size=9, bold=True, color="475569")
    font_kpi_value = Font(name="Segoe UI", size=14, bold=True, color="0F172A")

    thin_border_side = Side(border_style="thin", color="E2E8F0")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    thick_bottom = Border(bottom=Side(border_style="medium", color="0EA5E9"))
    total_border = Border(top=Side(border_style="thin", color="94A3B8"), bottom=Side(border_style="double", color="1E293B"))

    currency_format = '₹ #,##0.00'
    pct_format = '0.0%'

    if effective_user_id is None:
        user_email_display = f"All Users (Platform System-Wide) [Audited by Admin: {user.email}]"
    elif effective_user_id != user.id:
        target_u = db.query(User).filter(User.id == effective_user_id).first()
        user_email_display = f"User: {target_u.email if target_u else effective_user_id} [Audited by Admin: {user.email}]"
    else:
        user_email_display = f"User: {user.email}"

    # If transactions_only: build standalone Transactions sheet only
    if transactions_only:
        ws_tx = wb.active
        ws_tx.title = "Transactions"
        ws_tx.views.sheetView[0].showGridLines = True

        ws_tx.merge_cells("A1:G1")
        ws_tx["A1"].value = f"BudgetBuddy — Transactions Ledger ({report_data.summary.period_label})"
        ws_tx["A1"].font = font_title
        ws_tx["A1"].fill = header_fill
        ws_tx["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws_tx.row_dimensions[1].height = 32

        ws_tx.merge_cells("A2:G2")
        ws_tx["A2"].value = f"{user_email_display}   |   Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_tx["A2"].font = font_subtitle
        ws_tx["A2"].fill = header_fill
        ws_tx["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws_tx.row_dimensions[2].height = 20

        tx_headers = ["# ID", "Date", "Type", "Category / Source", "Account", "Description", "Amount (INR)"]
        for col_idx, h in enumerate(tx_headers, start=1):
            cell = ws_tx.cell(row=4, column=col_idx, value=h)
            cell.font = font_tbl_header
            cell.fill = subheader_fill
            cell.alignment = Alignment(horizontal="center" if col_idx in (1, 2, 3) else ("right" if col_idx == 7 else "left"), vertical="center")
            cell.border = thin_border
        ws_tx.row_dimensions[4].height = 24

        row_num = 5
        for item in report_data.transactions:
            is_inc = item.type == "income"
            ws_tx.cell(row=row_num, column=1, value=item.id).alignment = Alignment(horizontal="center")
            ws_tx.cell(row=row_num, column=2, value=item.date.strftime("%Y-%m-%d %H:%M") if item.date else "-").alignment = Alignment(horizontal="center")
            c3 = ws_tx.cell(row=row_num, column=3, value=item.type.upper())
            c3.alignment = Alignment(horizontal="center")
            c3.font = font_income if is_inc else font_expense
            ws_tx.cell(row=row_num, column=4, value=item.category_or_source).font = font_data_bold
            ws_tx.cell(row=row_num, column=5, value=item.account).font = font_data
            ws_tx.cell(row=row_num, column=6, value=item.description or "-").font = font_data
            c7 = ws_tx.cell(row=row_num, column=7, value=item.amount if is_inc else -item.amount)
            c7.font = font_income if is_inc else font_expense
            c7.number_format = currency_format
            c7.alignment = Alignment(horizontal="right")

            for c in range(1, 8):
                ws_tx.cell(row=row_num, column=c).border = thin_border
                if row_num % 2 == 0:
                    ws_tx.cell(row=row_num, column=c).fill = zebra_fill
            row_num += 1

        for col in ws_tx.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_tx.column_dimensions[col_letter].width = max(max_len + 4, 14)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1: Financial Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws_summary = wb.active
    ws_summary.title = "Financial Summary"
    ws_summary.views.sheetView[0].showGridLines = True

    # Main Brand Header Banner (A1:F2)
    ws_summary.merge_cells("A1:F1")
    cell_title = ws_summary["A1"]
    cell_title.value = "BudgetBuddy — Financial Report"
    cell_title.font = font_title
    cell_title.fill = header_fill
    cell_title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_summary.row_dimensions[1].height = 36

    ws_summary.merge_cells("A2:F2")
    cell_sub = ws_summary["A2"]
    generated_at_str = datetime.now().strftime("%d %B %Y, %I:%M %p")
    tier_note = "   |   [Basic Export: Preview (10 txns max). Upgrade to Premium for 4-Sheet Full Workbook]" if is_limited else ""
    cell_sub.value = f"Period: {report_data.summary.period_label}   |   Account User: {user_email_display}   |   Generated: {generated_at_str}{tier_note}"
    cell_sub.font = font_subtitle
    cell_sub.fill = header_fill
    cell_sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_summary.row_dimensions[2].height = 20

    # KPI Summary Cards Block (Row 4 to 6)
    kpis = [
        ("TOTAL INCOME", f"₹ {report_data.summary.total_income:,.2f}", kpi_income_fill, font_income, "B4:B5"),
        ("TOTAL EXPENSES", f"₹ {report_data.summary.total_expenses:,.2f}", kpi_expense_fill, font_expense, "C4:C5"),
        ("NET SAVINGS", f"₹ {report_data.summary.net_savings:,.2f}", kpi_net_fill, font_section, "D4:D5"),
        ("SAVINGS RATE", f"{report_data.summary.savings_rate}%", kpi_net_fill, font_section, "E4:E5"),
        ("TRANSACTIONS", f"{report_data.summary.total_transactions_count}", total_row_fill, font_section, "F4:F5"),
    ]

    for col_idx, (label, val, fill_color, val_font, _) in enumerate(kpis, start=2):
        col_letter = get_column_letter(col_idx)
        
        lbl_cell = ws_summary[f"{col_letter}4"]
        lbl_cell.value = label
        lbl_cell.font = font_kpi_label
        lbl_cell.fill = fill_color
        lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
        lbl_cell.border = thin_border

        val_cell = ws_summary[f"{col_letter}5"]
        val_cell.value = val
        val_cell.font = font_kpi_value
        val_cell.fill = fill_color
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        val_cell.border = thin_border

    ws_summary.row_dimensions[4].height = 18
    ws_summary.row_dimensions[5].height = 26

    # Section 1: Expense Breakdown Table
    curr_row = 8
    ws_summary.cell(row=curr_row, column=1, value="Expense Category Breakdown").font = font_section
    curr_row += 1

    cat_headers = ["Category", "Txn Count", "Total Amount (INR)", "% of Expenses"]
    for col_idx, h in enumerate(cat_headers, start=1):
        cell = ws_summary.cell(row=curr_row, column=col_idx, value=h)
        cell.font = font_tbl_header
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal="center" if col_idx in (2, 4) else ("right" if col_idx == 3 else "left"), vertical="center")
        cell.border = thin_border
    ws_summary.row_dimensions[curr_row].height = 22

    start_exp_row = curr_row + 1
    if not report_data.category_breakdown:
        curr_row += 1
        ws_summary.cell(row=curr_row, column=1, value="No expense records for this period.").font = font_data
    else:
        for item in report_data.category_breakdown:
            curr_row += 1
            ws_summary.cell(row=curr_row, column=1, value=item.category).font = font_data
            ws_summary.cell(row=curr_row, column=2, value=item.count).font = font_data
            ws_summary.cell(row=curr_row, column=2).alignment = Alignment(horizontal="center")
            
            amt_cell = ws_summary.cell(row=curr_row, column=3, value=item.amount)
            amt_cell.font = font_data
            amt_cell.number_format = currency_format
            amt_cell.alignment = Alignment(horizontal="right")

            pct_cell = ws_summary.cell(row=curr_row, column=4, value=f"{item.percentage}%")
            pct_cell.font = font_data
            pct_cell.alignment = Alignment(horizontal="center")

            for c in range(1, 5):
                ws_summary.cell(row=curr_row, column=c).border = thin_border
                if curr_row % 2 == 0:
                    ws_summary.cell(row=curr_row, column=c).fill = zebra_fill

    # Expense Total Row
    curr_row += 1
    t_lbl = ws_summary.cell(row=curr_row, column=1, value="Total Expenses")
    t_lbl.font = font_data_bold
    t_lbl.border = total_border
    
    t_cnt = ws_summary.cell(row=curr_row, column=2, value=report_data.summary.expense_count)
    t_cnt.font = font_data_bold
    t_cnt.alignment = Alignment(horizontal="center")
    t_cnt.border = total_border

    t_amt = ws_summary.cell(row=curr_row, column=3, value=report_data.summary.total_expenses)
    t_amt.font = font_expense
    t_amt.number_format = currency_format
    t_amt.border = total_border

    t_pct = ws_summary.cell(row=curr_row, column=4, value="100.0%")
    t_pct.font = font_data_bold
    t_pct.alignment = Alignment(horizontal="center")
    t_pct.border = total_border

    # Section 2: Income Source Breakdown Table
    curr_row += 3
    ws_summary.cell(row=curr_row, column=1, value="Income Source Breakdown").font = font_section
    curr_row += 1

    src_headers = ["Income Source", "Txn Count", "Total Amount (INR)", "% of Income"]
    for col_idx, h in enumerate(src_headers, start=1):
        cell = ws_summary.cell(row=curr_row, column=col_idx, value=h)
        cell.font = font_tbl_header
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal="center" if col_idx in (2, 4) else ("right" if col_idx == 3 else "left"), vertical="center")
        cell.border = thin_border
    ws_summary.row_dimensions[curr_row].height = 22

    if not report_data.source_breakdown:
        curr_row += 1
        ws_summary.cell(row=curr_row, column=1, value="No income records for this period.").font = font_data
    else:
        for item in report_data.source_breakdown:
            curr_row += 1
            ws_summary.cell(row=curr_row, column=1, value=item.source).font = font_data
            ws_summary.cell(row=curr_row, column=2, value=item.count).font = font_data
            ws_summary.cell(row=curr_row, column=2).alignment = Alignment(horizontal="center")
            
            amt_cell = ws_summary.cell(row=curr_row, column=3, value=item.amount)
            amt_cell.font = font_data
            amt_cell.number_format = currency_format
            amt_cell.alignment = Alignment(horizontal="right")

            pct_cell = ws_summary.cell(row=curr_row, column=4, value=f"{item.percentage}%")
            pct_cell.font = font_data
            pct_cell.alignment = Alignment(horizontal="center")

            for c in range(1, 5):
                ws_summary.cell(row=curr_row, column=c).border = thin_border
                if curr_row % 2 == 0:
                    ws_summary.cell(row=curr_row, column=c).fill = zebra_fill

    # Income Total Row
    curr_row += 1
    ti_lbl = ws_summary.cell(row=curr_row, column=1, value="Total Income")
    ti_lbl.font = font_data_bold
    ti_lbl.border = total_border
    
    ti_cnt = ws_summary.cell(row=curr_row, column=2, value=report_data.summary.income_count)
    ti_cnt.font = font_data_bold
    ti_cnt.alignment = Alignment(horizontal="center")
    ti_cnt.border = total_border

    ti_amt = ws_summary.cell(row=curr_row, column=3, value=report_data.summary.total_income)
    ti_amt.font = font_income
    ti_amt.number_format = currency_format
    ti_amt.border = total_border

    ti_pct = ws_summary.cell(row=curr_row, column=4, value="100.0%")
    ti_pct.font = font_data_bold
    ti_pct.alignment = Alignment(horizontal="center")
    ti_pct.border = total_border

    # Section 3: Account Cash Flow Table
    curr_row += 3
    ws_summary.cell(row=curr_row, column=1, value="Account Cash Flow Breakdown").font = font_section
    curr_row += 1

    acct_headers = ["Account Name", "Transactions", "Total Inflow (INR)", "Total Outflow (INR)", "Net Flow (INR)"]
    for col_idx, h in enumerate(acct_headers, start=1):
        cell = ws_summary.cell(row=curr_row, column=col_idx, value=h)
        cell.font = font_tbl_header
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal="center" if col_idx == 2 else ("right" if col_idx >= 3 else "left"), vertical="center")
        cell.border = thin_border
    ws_summary.row_dimensions[curr_row].height = 22

    if not report_data.account_breakdown:
        curr_row += 1
        ws_summary.cell(row=curr_row, column=1, value="No account activity found for this period.").font = font_data
    else:
        for item in report_data.account_breakdown:
            curr_row += 1
            ws_summary.cell(row=curr_row, column=1, value=item.account).font = font_data
            ws_summary.cell(row=curr_row, column=2, value=item.transaction_count).font = font_data
            ws_summary.cell(row=curr_row, column=2).alignment = Alignment(horizontal="center")
            
            in_cell = ws_summary.cell(row=curr_row, column=3, value=item.income_amount)
            in_cell.font = font_income
            in_cell.number_format = currency_format
            in_cell.alignment = Alignment(horizontal="right")

            out_cell = ws_summary.cell(row=curr_row, column=4, value=item.expense_amount)
            out_cell.font = font_expense
            out_cell.number_format = currency_format
            out_cell.alignment = Alignment(horizontal="right")

            net_cell = ws_summary.cell(row=curr_row, column=5, value=item.net_amount)
            net_cell.font = font_income if item.net_amount >= 0 else font_expense
            net_cell.number_format = currency_format
            net_cell.alignment = Alignment(horizontal="right")

            for c in range(1, 6):
                ws_summary.cell(row=curr_row, column=c).border = thin_border
                if curr_row % 2 == 0:
                    ws_summary.cell(row=curr_row, column=c).fill = zebra_fill


    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2: All Transactions
    # ══════════════════════════════════════════════════════════════════════════
    ws_all = wb.create_sheet(title="All Transactions")
    ws_all.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_all.merge_cells("A1:G1")
    ws_all["A1"].value = f"All Transactions Ledger — {report_data.summary.period_label}"
    ws_all["A1"].font = font_title
    ws_all["A1"].fill = header_fill
    ws_all["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_all.row_dimensions[1].height = 32

    # Headers
    tx_headers = ["# ID", "Date", "Type", "Category / Source", "Account", "Description", "Amount (INR)"]
    for col_idx, h in enumerate(tx_headers, start=1):
        cell = ws_all.cell(row=3, column=col_idx, value=h)
        cell.font = font_tbl_header
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal="center" if col_idx in (1, 2, 3) else ("right" if col_idx == 7 else "left"), vertical="center")
        cell.border = thin_border
    ws_all.row_dimensions[3].height = 24

    row_num = 4
    for item in report_data.transactions:
        is_inc = item.type == "income"
        
        c1 = ws_all.cell(row=row_num, column=1, value=item.id)
        c1.alignment = Alignment(horizontal="center")
        c1.font = font_data

        c2 = ws_all.cell(row=row_num, column=2, value=item.date.strftime("%Y-%m-%d %H:%M") if item.date else "-")
        c2.alignment = Alignment(horizontal="center")
        c2.font = font_data

        c3 = ws_all.cell(row=row_num, column=3, value=item.type.upper())
        c3.alignment = Alignment(horizontal="center")
        c3.font = font_income if is_inc else font_expense

        c4 = ws_all.cell(row=row_num, column=4, value=item.category_or_source)
        c4.font = font_data_bold

        c5 = ws_all.cell(row=row_num, column=5, value=item.account)
        c5.font = font_data

        c6 = ws_all.cell(row=row_num, column=6, value=item.description or "-")
        c6.font = font_data

        c7 = ws_all.cell(row=row_num, column=7, value=item.amount if is_inc else -item.amount)
        c7.font = font_income if is_inc else font_expense
        c7.number_format = currency_format
        c7.alignment = Alignment(horizontal="right")

        for c in range(1, 8):
            ws_all.cell(row=row_num, column=c).border = thin_border
            if row_num % 2 == 0:
                ws_all.cell(row=row_num, column=c).fill = zebra_fill

        row_num += 1

    # Total Summary Line
    if report_data.transactions:
        ws_all.cell(row=row_num, column=1, value="").border = total_border
        ws_all.cell(row=row_num, column=2, value="").border = total_border
        ws_all.cell(row=row_num, column=3, value="").border = total_border
        ws_all.cell(row=row_num, column=4, value="").border = total_border
        ws_all.cell(row=row_num, column=5, value="").border = total_border
        
        tot_lbl = ws_all.cell(row=row_num, column=6, value="Net Total:")
        tot_lbl.font = font_data_bold
        tot_lbl.alignment = Alignment(horizontal="right")
        tot_lbl.border = total_border

        tot_val = ws_all.cell(row=row_num, column=7, value=report_data.summary.net_savings)
        tot_val.font = font_income if report_data.summary.net_savings >= 0 else font_expense
        tot_val.number_format = currency_format
        tot_val.alignment = Alignment(horizontal="right")
        tot_val.border = total_border


    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3: Income Details
    # ══════════════════════════════════════════════════════════════════════════
    ws_inc = wb.create_sheet(title="Income Details")
    ws_inc.views.sheetView[0].showGridLines = True

    ws_inc.merge_cells("A1:E1")
    ws_inc["A1"].value = f"Income Transactions — {report_data.summary.period_label}"
    ws_inc["A1"].font = font_title
    ws_inc["A1"].fill = header_fill
    ws_inc["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_inc.row_dimensions[1].height = 32

    inc_headers = ["# ID", "Date", "Source", "Account", "Amount (INR)"]
    for col_idx, h in enumerate(inc_headers, start=1):
        cell = ws_inc.cell(row=3, column=col_idx, value=h)
        cell.font = font_tbl_header
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal="center" if col_idx in (1, 2) else ("right" if col_idx == 5 else "left"), vertical="center")
        cell.border = thin_border
    ws_inc.row_dimensions[3].height = 24

    r_inc = 4
    income_txns = [t for t in report_data.transactions if t.type == "income"]
    for item in income_txns:
        ws_inc.cell(row=r_inc, column=1, value=item.id).alignment = Alignment(horizontal="center")
        ws_inc.cell(row=r_inc, column=2, value=item.date.strftime("%Y-%m-%d %H:%M") if item.date else "-").alignment = Alignment(horizontal="center")
        ws_inc.cell(row=r_inc, column=3, value=item.category_or_source).font = font_data_bold
        ws_inc.cell(row=r_inc, column=4, value=item.account)
        
        amt_cell = ws_inc.cell(row=r_inc, column=5, value=item.amount)
        amt_cell.font = font_income
        amt_cell.number_format = currency_format
        amt_cell.alignment = Alignment(horizontal="right")

        for c in range(1, 6):
            ws_inc.cell(row=r_inc, column=c).border = thin_border
            if r_inc % 2 == 0:
                ws_inc.cell(row=r_inc, column=c).fill = zebra_fill
        r_inc += 1

    # Total Income Row
    if income_txns:
        for c in range(1, 4):
            ws_inc.cell(row=r_inc, column=c).border = total_border
        t_lbl = ws_inc.cell(row=r_inc, column=4, value="Total Income:")
        t_lbl.font = font_data_bold
        t_lbl.alignment = Alignment(horizontal="right")
        t_lbl.border = total_border

        t_val = ws_inc.cell(row=r_inc, column=5, value=report_data.summary.total_income)
        t_val.font = font_income
        t_val.number_format = currency_format
        t_val.border = total_border


    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4: Expense Details
    # ══════════════════════════════════════════════════════════════════════════
    ws_exp = wb.create_sheet(title="Expense Details")
    ws_exp.views.sheetView[0].showGridLines = True

    ws_exp.merge_cells("A1:F1")
    ws_exp["A1"].value = f"Expense Transactions — {report_data.summary.period_label}"
    ws_exp["A1"].font = font_title
    ws_exp["A1"].fill = header_fill
    ws_exp["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_exp.row_dimensions[1].height = 32

    exp_headers = ["# ID", "Date", "Category", "Account", "Description", "Amount (INR)"]
    for col_idx, h in enumerate(exp_headers, start=1):
        cell = ws_exp.cell(row=3, column=col_idx, value=h)
        cell.font = font_tbl_header
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal="center" if col_idx in (1, 2) else ("right" if col_idx == 6 else "left"), vertical="center")
        cell.border = thin_border
    ws_exp.row_dimensions[3].height = 24

    r_exp = 4
    expense_txns = [t for t in report_data.transactions if t.type == "expense"]
    for item in expense_txns:
        ws_exp.cell(row=r_exp, column=1, value=item.id).alignment = Alignment(horizontal="center")
        ws_exp.cell(row=r_exp, column=2, value=item.date.strftime("%Y-%m-%d %H:%M") if item.date else "-").alignment = Alignment(horizontal="center")
        ws_exp.cell(row=r_exp, column=3, value=item.category_or_source).font = font_data_bold
        ws_exp.cell(row=r_exp, column=4, value=item.account)
        ws_exp.cell(row=r_exp, column=5, value=item.description or "-")
        
        amt_cell = ws_exp.cell(row=r_exp, column=6, value=item.amount)
        amt_cell.font = font_expense
        amt_cell.number_format = currency_format
        amt_cell.alignment = Alignment(horizontal="right")

        for c in range(1, 7):
            ws_exp.cell(row=r_exp, column=c).border = thin_border
            if r_exp % 2 == 0:
                ws_exp.cell(row=r_exp, column=c).fill = zebra_fill
        r_exp += 1

    # Total Expense Row
    if expense_txns:
        for c in range(1, 5):
            ws_exp.cell(row=r_exp, column=c).border = total_border
        t_lbl = ws_exp.cell(row=r_exp, column=5, value="Total Expenses:")
        t_lbl.font = font_data_bold
        t_lbl.alignment = Alignment(horizontal="right")
        t_lbl.border = total_border

        t_val = ws_exp.cell(row=r_exp, column=6, value=report_data.summary.total_expenses)
        t_val.font = font_expense
        t_val.number_format = currency_format
        t_val.border = total_border

    # ══════════════════════════════════════════════════════════════════════════
    # Column Width Auto-Fitting for all sheets
    # ══════════════════════════════════════════════════════════════════════════
    for ws in [ws_summary, ws_all, ws_inc, ws_exp]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Ignore merged row 1 & 2 for width calculation
                if cell.row in (1, 2) and col_letter in ("A", "B", "C", "D", "E", "F", "G"):
                    continue
                if cell.value:
                    val_str = str(cell.value)
                    max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


class NumberedCanvas(canvas.Canvas):
    """
    Two-pass canvas to dynamically compute and print total page numbers (e.g., 'Page 1 of 4')
    and professional running headers and footers on multi-page reports.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []
        self.report_period_label = ""
        self.user_email = ""

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count: int):
        self.saveState()
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#64748B"))

        # Footer
        footer_left = "BudgetBuddy Financial Performance & Audit Report • Confidential"
        page_str = f"Page {self._pageNumber} of {page_count}"
        self.drawString(36, 22, footer_left)
        self.drawRightString(A4[0] - 36, 22, page_str)

        # Footer divider line
        self.setStrokeColor(colors.HexColor("#E2E8F0"))
        self.setLineWidth(0.5)
        self.line(36, 32, A4[0] - 36, 32)

        # Running Header (pages 2+)
        if self._pageNumber > 1:
            self.drawString(36, A4[1] - 25, "BudgetBuddy — Financial & Transaction Report")
            if self.report_period_label:
                self.drawRightString(A4[0] - 36, A4[1] - 25, self.report_period_label)
            self.setStrokeColor(colors.HexColor("#CBD5E1"))
            self.setLineWidth(0.5)
            self.line(36, A4[1] - 30, A4[0] - 36, A4[1] - 30)

        self.restoreState()


def create_numbered_canvas(period_label: str, user_email: str):
    class CustomNumberedCanvas(NumberedCanvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self.report_period_label = period_label
            self.user_email = user_email
    return CustomNumberedCanvas


def generate_pdf_report(
    db: Session,
    user: User,
    period_type: str = "month",
    month: Optional[int] = None,
    year: Optional[int] = None,
    start_date_str: Optional[str] = None,
    end_date_str: Optional[str] = None,
    transaction_type: str = "all",
    category: Optional[str] = None,
    account: Optional[str] = None,
    is_limited: bool = False,
    target_user_id: Optional[int] = None,
) -> io.BytesIO:
    """
    Generates a high-quality, professional PDF financial report including:
    - Executive Branding & Period Metadata
    - KPI Metrics Summary (Total Income, Total Expenses, Net Savings, Savings Rate, Transaction Counts)
    - Expense Category Breakdown Table
    - Income Source Breakdown Table
    - Account Cash Flow Breakdown Table
    - Full Transaction History Ledger with individual line items and amounts
    """
    effective_user_id = target_user_id if target_user_id is not None else (user.id if user.role != "admin" else None)
    report_data = get_user_report_data(
        db=db,
        user_id=effective_user_id,
        period_type=period_type,
        month=month,
        year=year,
        start_date_str=start_date_str,
        end_date_str=end_date_str,
        transaction_type=transaction_type,
        category=category,
        account=account,
        is_limited=is_limited,
        is_admin=(user.role == "admin"),
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=36,
        rightMargin=36,
        topMargin=40,
        bottomMargin=42,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "DocTitle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=15,
        leading=18,
        textColor=colors.white,
    )
    sub_style = ParagraphStyle(
        "DocSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor("#94A3B8"),
    )
    meta_style = ParagraphStyle(
        "DocMeta",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=11,
        alignment=TA_RIGHT,
        textColor=colors.HexColor("#E2E8F0"),
    )
    sec_head = ParagraphStyle(
        "SecHead",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10.5,
        leading=13,
        textColor=colors.HexColor("#0F172A"),
        spaceAfter=3,
    )
    tbl_hdr = ParagraphStyle(
        "TblHdr",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=colors.white,
        alignment=TA_LEFT,
    )
    tbl_hdr_r = ParagraphStyle(
        "TblHdrR",
        parent=tbl_hdr,
        alignment=TA_RIGHT,
    )
    tbl_hdr_c = ParagraphStyle(
        "TblHdrC",
        parent=tbl_hdr,
        alignment=TA_CENTER,
    )
    cell_txt = ParagraphStyle(
        "CellTxt",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#1E293B"),
    )
    cell_txt_b = ParagraphStyle(
        "CellTxtB",
        parent=cell_txt,
        fontName="Helvetica-Bold",
    )
    cell_txt_r = ParagraphStyle(
        "CellTxtR",
        parent=cell_txt,
        alignment=TA_RIGHT,
    )
    cell_txt_c = ParagraphStyle(
        "CellTxtC",
        parent=cell_txt,
        alignment=TA_CENTER,
    )
    cell_inc = ParagraphStyle(
        "CellInc",
        parent=cell_txt_r,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#15803D"),
    )
    cell_exp = ParagraphStyle(
        "CellExp",
        parent=cell_txt_r,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#BE123C"),
    )

    story = []
    page_w = 523.27

    # 1. Header Banner
    now_str = datetime.now().strftime("%d %b %Y, %I:%M %p")
    if effective_user_id is None:
        user_email_display = f"All Users (Platform System-Wide) | Admin: {user.email}"
    elif effective_user_id != user.id:
        target_u = db.query(User).filter(User.id == effective_user_id).first()
        user_email_display = f"User: {target_u.email if target_u else effective_user_id} | Admin: {user.email}"
    else:
        user_email_display = user.email if user else "User"

    banner_data = [
        [
            [
                Paragraph("BudgetBuddy", title_style),
                Spacer(1, 2),
                Paragraph("Financial Performance & Audit Report", sub_style),
            ],
            [
                Paragraph(f"<b>Period:</b> {report_data.summary.period_label}", meta_style),
                Paragraph(f"<b>Account:</b> {user_email_display}", meta_style),
                Paragraph(f"<b>Generated:</b> {now_str}", meta_style),
                Paragraph("<b>Tier:</b> Basic (10 txns preview)", meta_style) if is_limited else Paragraph("<b>Tier:</b> Premium Audit", meta_style),
            ],
        ]
    ]
    banner_table = Table(banner_data, colWidths=[280, page_w - 280])
    banner_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0F172A")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
    ]))
    story.append(banner_table)
    story.append(Spacer(1, 8))

    # 2. Executive KPI Cards
    kpi_w = page_w / 5.0
    kpi_data = [
        [
            Paragraph('<font color="#15803D" size="6.5"><b>TOTAL INCOME</b></font>', cell_txt_c),
            Paragraph('<font color="#BE123C" size="6.5"><b>TOTAL EXPENSES</b></font>', cell_txt_c),
            Paragraph('<font color="#0369A1" size="6.5"><b>NET SAVINGS</b></font>', cell_txt_c),
            Paragraph('<font color="#7E22CE" size="6.5"><b>SAVINGS RATE</b></font>', cell_txt_c),
            Paragraph('<font color="#334155" size="6.5"><b>TRANSACTIONS</b></font>', cell_txt_c),
        ],
        [
            Paragraph(f'<font color="#15803D" size="9.5"><b>INR {report_data.summary.total_income:,.2f}</b></font>', cell_txt_c),
            Paragraph(f'<font color="#BE123C" size="9.5"><b>INR {report_data.summary.total_expenses:,.2f}</b></font>', cell_txt_c),
            Paragraph(f'<font color="#0369A1" size="9.5"><b>INR {report_data.summary.net_savings:,.2f}</b></font>', cell_txt_c),
            Paragraph(f'<font color="#7E22CE" size="9.5"><b>{report_data.summary.savings_rate}%</b></font>', cell_txt_c),
            Paragraph(f'<font color="#0F172A" size="9.5"><b>{report_data.summary.total_transactions_count}</b></font>', cell_txt_c),
        ]
    ]
    kpi_table = Table(kpi_data, colWidths=[kpi_w]*5)
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 1), colors.HexColor("#F0FDF4")),
        ("BACKGROUND", (1, 0), (1, 1), colors.HexColor("#FFF1F2")),
        ("BACKGROUND", (2, 0), (2, 1), colors.HexColor("#F0F9FF")),
        ("BACKGROUND", (3, 0), (3, 1), colors.HexColor("#FAF5FF")),
        ("BACKGROUND", (4, 0), (4, 1), colors.HexColor("#F8FAFC")),
        ("BOX", (0, 0), (0, 1), 0.5, colors.HexColor("#86EFAC")),
        ("BOX", (1, 0), (1, 1), 0.5, colors.HexColor("#FECDD3")),
        ("BOX", (2, 0), (2, 1), 0.5, colors.HexColor("#BAE6FD")),
        ("BOX", (3, 0), (3, 1), 0.5, colors.HexColor("#E9D5FF")),
        ("BOX", (4, 0), (4, 1), 0.5, colors.HexColor("#CBD5E1")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 10))

    # 3. Expense Category Breakdown Table
    story.append(Paragraph("Expense Category Breakdown", sec_head))
    exp_table_data = [
        [
            Paragraph("Category", tbl_hdr),
            Paragraph("Txn Count", tbl_hdr_c),
            Paragraph("Total Amount", tbl_hdr_r),
            Paragraph("% of Expenses", tbl_hdr_c),
        ]
    ]
    if not report_data.category_breakdown:
        exp_table_data.append([
            Paragraph("No expense records for this period.", cell_txt),
            Paragraph("-", cell_txt_c),
            Paragraph("INR 0.00", cell_txt_r),
            Paragraph("0.0%", cell_txt_c),
        ])
    else:
        for item in report_data.category_breakdown:
            exp_table_data.append([
                Paragraph(item.category, cell_txt),
                Paragraph(str(item.count), cell_txt_c),
                Paragraph(f"INR {item.amount:,.2f}", cell_txt_r),
                Paragraph(f"{item.percentage:.1f}%", cell_txt_c),
            ])
    # Expense Total Row
    exp_table_data.append([
        Paragraph("Total Expenses", cell_txt_b),
        Paragraph(str(report_data.summary.expense_count), cell_txt_c),
        Paragraph(f"INR {report_data.summary.total_expenses:,.2f}", cell_exp),
        Paragraph("100.0%", cell_txt_c),
    ])

    col_w_4 = [170, 70, 140, 143.27]
    exp_tbl = Table(exp_table_data, colWidths=col_w_4)
    exp_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
        ("TOPPADDING", (0, 0), (-1, -1), 3.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("LINEBELOW", (0, -1), (-1, -1), 1.2, colors.HexColor("#1E293B")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F1F5F9")),
    ]
    for r_idx in range(1, len(exp_table_data) - 1):
        if r_idx % 2 == 0:
            exp_style.append(("BACKGROUND", (0, r_idx), (-1, r_idx), colors.HexColor("#F8FAFC")))
    exp_tbl.setStyle(TableStyle(exp_style))
    story.append(exp_tbl)
    story.append(Spacer(1, 10))

    # 4. Income Source Breakdown Table
    story.append(Paragraph("Income Source Breakdown", sec_head))
    inc_table_data = [
        [
            Paragraph("Income Source", tbl_hdr),
            Paragraph("Txn Count", tbl_hdr_c),
            Paragraph("Total Amount", tbl_hdr_r),
            Paragraph("% of Income", tbl_hdr_c),
        ]
    ]
    if not report_data.source_breakdown:
        inc_table_data.append([
            Paragraph("No income records for this period.", cell_txt),
            Paragraph("-", cell_txt_c),
            Paragraph("INR 0.00", cell_txt_r),
            Paragraph("0.0%", cell_txt_c),
        ])
    else:
        for item in report_data.source_breakdown:
            inc_table_data.append([
                Paragraph(item.source, cell_txt),
                Paragraph(str(item.count), cell_txt_c),
                Paragraph(f"INR {item.amount:,.2f}", cell_txt_r),
                Paragraph(f"{item.percentage:.1f}%", cell_txt_c),
            ])
    # Income Total Row
    inc_table_data.append([
        Paragraph("Total Income", cell_txt_b),
        Paragraph(str(report_data.summary.income_count), cell_txt_c),
        Paragraph(f"INR {report_data.summary.total_income:,.2f}", cell_inc),
        Paragraph("100.0%", cell_txt_c),
    ])

    inc_tbl = Table(inc_table_data, colWidths=col_w_4)
    inc_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
        ("TOPPADDING", (0, 0), (-1, -1), 3.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("LINEBELOW", (0, -1), (-1, -1), 1.2, colors.HexColor("#1E293B")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F1F5F9")),
    ]
    for r_idx in range(1, len(inc_table_data) - 1):
        if r_idx % 2 == 0:
            inc_style.append(("BACKGROUND", (0, r_idx), (-1, r_idx), colors.HexColor("#F8FAFC")))
    inc_tbl.setStyle(TableStyle(inc_style))
    story.append(inc_tbl)
    story.append(Spacer(1, 10))

    # 5. Account Cash Flow Breakdown Table
    story.append(Paragraph("Account Cash Flow Breakdown", sec_head))
    acct_table_data = [
        [
            Paragraph("Account Name", tbl_hdr),
            Paragraph("Txns", tbl_hdr_c),
            Paragraph("Inflow", tbl_hdr_r),
            Paragraph("Outflow", tbl_hdr_r),
            Paragraph("Net Flow", tbl_hdr_r),
        ]
    ]
    if not report_data.account_breakdown:
        acct_table_data.append([
            Paragraph("No account activity found.", cell_txt),
            Paragraph("-", cell_txt_c),
            Paragraph("INR 0.00", cell_txt_r),
            Paragraph("INR 0.00", cell_txt_r),
            Paragraph("INR 0.00", cell_txt_r),
        ])
    else:
        for item in report_data.account_breakdown:
            net_style = cell_inc if item.net_amount >= 0 else cell_exp
            acct_table_data.append([
                Paragraph(item.account, cell_txt),
                Paragraph(str(item.transaction_count), cell_txt_c),
                Paragraph(f"INR {item.income_amount:,.2f}", cell_inc),
                Paragraph(f"INR {item.expense_amount:,.2f}", cell_exp),
                Paragraph(f"{'+' if item.net_amount >= 0 else ''}INR {item.net_amount:,.2f}", net_style),
            ])

    acct_col_w = [143.27, 50, 110, 110, 110]
    acct_tbl = Table(acct_table_data, colWidths=acct_col_w)
    acct_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
        ("TOPPADDING", (0, 0), (-1, -1), 3.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
    ]
    for r_idx in range(1, len(acct_table_data)):
        if r_idx % 2 == 0:
            acct_style.append(("BACKGROUND", (0, r_idx), (-1, r_idx), colors.HexColor("#F8FAFC")))
    acct_tbl.setStyle(TableStyle(acct_style))
    story.append(acct_tbl)
    story.append(Spacer(1, 14))

    # 6. Detailed Transactions Ledger
    story.append(Paragraph("Transaction History Ledger", sec_head))
    tx_table_data = [
        [
            Paragraph("# ID", tbl_hdr_c),
            Paragraph("Date", tbl_hdr_c),
            Paragraph("Type", tbl_hdr_c),
            Paragraph("Category / Source", tbl_hdr),
            Paragraph("Account", tbl_hdr),
            Paragraph("Description", tbl_hdr),
            Paragraph("Amount", tbl_hdr_r),
        ]
    ]

    tx_col_w = [32, 65, 48, 92, 65, 136.27, 85]

    if not report_data.transactions:
        tx_table_data.append([
            Paragraph("-", cell_txt_c),
            Paragraph("-", cell_txt_c),
            Paragraph("-", cell_txt_c),
            Paragraph("No transactions matching filters.", cell_txt),
            Paragraph("-", cell_txt),
            Paragraph("-", cell_txt),
            Paragraph("INR 0.00", cell_txt_r),
        ])
    else:
        for tx in report_data.transactions:
            is_inc = tx.type == "income"
            date_str = tx.date.strftime("%Y-%m-%d %H:%M") if tx.date else "-"
            type_label = "INCOME" if is_inc else "EXPENSE"
            type_style = cell_inc if is_inc else cell_exp
            amt_str = f"{'+' if is_inc else '-'}INR {tx.amount:,.2f}"

            tx_table_data.append([
                Paragraph(f"#{tx.id}", cell_txt_c),
                Paragraph(date_str, cell_txt_c),
                Paragraph(type_label, type_style),
                Paragraph(tx.category_or_source or "-", cell_txt_b),
                Paragraph(tx.account or "Cash", cell_txt),
                Paragraph(tx.description or "-", cell_txt),
                Paragraph(amt_str, type_style),
            ])

    # Summary row at bottom of ledger
    tx_table_data.append([
        Paragraph("", cell_txt),
        Paragraph("", cell_txt),
        Paragraph("", cell_txt),
        Paragraph("", cell_txt),
        Paragraph("", cell_txt),
        Paragraph("Net Ledger Total:", cell_txt_b),
        Paragraph(
            f"{'+' if report_data.summary.net_savings >= 0 else ''}INR {report_data.summary.net_savings:,.2f}",
            cell_inc if report_data.summary.net_savings >= 0 else cell_exp,
        ),
    ])

    tx_tbl = Table(tx_table_data, colWidths=tx_col_w, repeatRows=1)
    tx_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("LINEBELOW", (0, -1), (-1, -1), 1.2, colors.HexColor("#1E293B")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F1F5F9")),
    ]
    for r_idx in range(1, len(tx_table_data) - 1):
        if r_idx % 2 == 0:
            tx_style.append(("BACKGROUND", (0, r_idx), (-1, r_idx), colors.HexColor("#F8FAFC")))
    tx_tbl.setStyle(TableStyle(tx_style))
    story.append(tx_tbl)

    # Build Document
    canvas_factory = create_numbered_canvas(
        report_data.summary.period_label,
        user_email_display,
    )
    doc.build(story, canvasmaker=canvas_factory)
    buf.seek(0)
    return buf


def generate_csv_report(
    db: Session,
    user_id: Optional[int] = None,
    period_type: str = "month",
    month: Optional[int] = None,
    year: Optional[int] = None,
    start_date_str: Optional[str] = None,
    end_date_str: Optional[str] = None,
    transaction_type: str = "all",
    category: Optional[str] = None,
    account: Optional[str] = None,
    requesting_user: Optional[User] = None,
) -> io.StringIO:
    """
    Generates a CSV transaction ledger file.
    """
    report_data = get_user_report_data(
        db=db,
        user_id=user_id,
        period_type=period_type,
        month=month,
        year=year,
        start_date_str=start_date_str,
        end_date_str=end_date_str,
        transaction_type=transaction_type,
        category=category,
        account=account,
        is_admin=(requesting_user.role == "admin") if requesting_user else False,
    )

    output = io.StringIO()
    writer = csv.writer(output)

    # Metadata Header
    writer.writerow(["# BudgetBuddy Transaction History Report"])
    writer.writerow(["# Period", report_data.summary.period_label])
    if user_id is None:
        writer.writerow(["# Scope", "All Users (Platform System-Wide)"])
        if requesting_user:
            writer.writerow(["# Audited By", requesting_user.email])
    else:
        target_u = db.query(User).filter(User.id == user_id).first()
        writer.writerow(["# Account", target_u.email if target_u else f"User #{user_id}"])
    writer.writerow(["# Total Income", report_data.summary.total_income])
    writer.writerow(["# Total Expenses", report_data.summary.total_expenses])
    writer.writerow(["# Net Savings", report_data.summary.net_savings])
    writer.writerow(["# Total Transactions", report_data.summary.total_transactions_count])
    writer.writerow([])

    # Table Header
    is_system_wide = (user_id is None)
    if is_system_wide:
        writer.writerow(["ID", "User Email", "Date", "Type", "Category/Source", "Account", "Description", "Amount"])
    else:
        writer.writerow(["ID", "Date", "Type", "Category/Source", "Account", "Description", "Amount"])

    for item in report_data.transactions:
        row = [item.id]
        if is_system_wide:
            row.append(item.user_email or "-")
        row.extend([
            item.date.strftime("%Y-%m-%d %H:%M") if item.date else "",
            item.type.upper(),
            item.category_or_source,
            item.account,
            item.description or "",
            item.amount if item.type == "income" else -item.amount,
        ])
        writer.writerow(row)

    output.seek(0)
    return output
