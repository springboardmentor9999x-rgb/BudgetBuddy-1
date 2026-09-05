from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from app.crud.reports import get_report
from app.models.profile import Profile
from app.models.user import User


# ============================================================
# COLORS
# ============================================================

NAVY = "071A2B"
NAVY_2 = "0B263D"
GREEN = "059669"
GREEN_LIGHT = "ECFDF5"
RED = "DC2626"
RED_LIGHT = "FEF2F2"
BLUE = "2563EB"
BLUE_LIGHT = "EFF6FF"
SLATE = "64748B"
LIGHT_SLATE = "F8FAFC"
BORDER = "CBD5E1"
WHITE = "FFFFFF"


# ============================================================
# HELPERS
# ============================================================

def date_text(value):
    if not value:
        return "—"

    return value.strftime("%d %b %Y")


def get_holder_name(db, user_id):
    user = (
        db.query(User)
        .filter(User.id == user_id)
        .first()
    )

    profile = (
        db.query(Profile)
        .filter(Profile.user_id == user_id)
        .first()
    )

    return (
        getattr(profile, "full_name", None)
        or getattr(profile, "name", None)
        or getattr(user, "full_name", None)
        or getattr(user, "name", None)
        or getattr(user, "email", None)
        or "Account Holder"
    )


# ============================================================
# EXCEL REPORT
# ============================================================

def generate_report_excel(
    db,
    user_id,
    period,
    start_date=None,
    end_date=None,
):
    report = get_report(
        db=db,
        user_id=user_id,
        period=period,
        start_date=start_date,
        end_date=end_date,
    )

    holder_name = get_holder_name(
        db,
        user_id,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Report"

    # ========================================================
    # PAGE SETTINGS
    # ========================================================

    ws.sheet_view.showGridLines = False


    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_margins = PageMargins(
        left=0.3,
        right=0.3,
        top=0.5,
        bottom=0.5,
        header=0.2,
        footer=0.2,
    )

    ws.print_title_rows = "1:15"

    # ========================================================
    # COLUMN WIDTHS
    # ========================================================

    widths = {
        "A": 24,
        "B": 18,
        "C": 23,
        "D": 42,
        "E": 20,
        "F": 20,
        "G": 25,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    # ========================================================
    # COMMON STYLES
    # ========================================================

    thin = Side(
        style="thin",
        color=BORDER,
    )

    medium = Side(
        style="medium",
        color=NAVY,
    )

    box_border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin,
    )

    # ========================================================
    # HEADER
    # ========================================================

    ws.merge_cells("A1:E2")

    title = ws["A1"]
    title.value = "BUDGET BUDDY"
    title.font = Font(
        name="Arial",
        size=24,
        bold=True,
        color=WHITE,
    )
    title.fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )
    title.alignment = Alignment(
        horizontal="left",
        vertical="center",
    )

    for row in ws["A1:E2"]:
        for cell in row:
            cell.fill = PatternFill(
                "solid",
                fgColor=NAVY,
            )

    ws.merge_cells("F1:G1")

    ws["F1"] = "ACCOUNT HOLDER"
    ws["F1"].font = Font(
        name="Arial",
        size=9,
        bold=True,
        color=WHITE,
    )
    ws["F1"].fill = PatternFill(
        "solid",
        fgColor=NAVY_2,
    )
    ws["F1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws.merge_cells("F2:G2")

    ws["F2"] = holder_name
    ws["F2"].font = Font(
        name="Arial",
        size=12,
        bold=True,
        color=NAVY,
    )
    ws["F2"].fill = PatternFill(
        "solid",
        fgColor=GREEN_LIGHT,
    )
    ws["F2"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 25

    # ========================================================
    # REPORT INFORMATION
    # ========================================================

    ws.merge_cells("A4:G4")

    ws["A4"] = f"{period.title()} Financial Report"
    ws["A4"].font = Font(
        name="Arial",
        size=19,
        bold=True,
        color=NAVY,
    )
    ws["A4"].alignment = Alignment(
        horizontal="left",
        vertical="center",
    )

    ws.merge_cells("A5:G5")

    ws["A5"] = (
        f"Period: {date_text(report['start_date'])}"
        f"  —  {date_text(report['end_date'])}"
    )

    ws["A5"].font = Font(
        name="Arial",
        size=10,
        color=SLATE,
    )

    ws["A5"].alignment = Alignment(
        horizontal="left",
    )

    # ========================================================
    # SUMMARY SECTION
    # ========================================================

    ws.merge_cells("A7:G7")

    ws["A7"] = "FINANCIAL SUMMARY"
    ws["A7"].font = Font(
        name="Arial",
        size=13,
        bold=True,
        color=WHITE,
    )
    ws["A7"].fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )
    ws["A7"].alignment = Alignment(
        horizontal="left",
        vertical="center",
    )

    summary = [
        (
            "TOTAL INCOME",
            report["summary"]["total_income"],
            GREEN_LIGHT,
            GREEN,
        ),
        (
            "TOTAL EXPENSES",
            report["summary"]["total_expenses"],
            RED_LIGHT,
            RED,
        ),
        (
            "NET BALANCE",
            report["summary"]["net_balance"],
            BLUE_LIGHT,
            BLUE,
        ),
        (
            "SAVINGS",
            report["summary"]["savings"],
            GREEN_LIGHT,
            GREEN,
        ),
    ]

    summary_columns = ["A", "C", "E", "G"]

    for index, (
        label,
        value,
        fill_color,
        font_color,
    ) in enumerate(summary):

        col = summary_columns[index]

        ws.merge_cells(
            f"{col}8:{col}8"
        )

        label_cell = ws[f"{col}8"]
        label_cell.value = label
        label_cell.font = Font(
            name="Arial",
            size=9,
            bold=True,
            color=font_color,
        )
        label_cell.fill = PatternFill(
            "solid",
            fgColor=fill_color,
        )
        label_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        label_cell.border = box_border

        ws.merge_cells(
            f"{col}9:{col}9"
        )

        value_cell = ws[f"{col}9"]
        value_cell.value = float(value or 0)
        value_cell.number_format = '₹#,##0.00'
        value_cell.font = Font(
            name="Arial",
            size=15,
            bold=True,
            color=font_color,
        )
        value_cell.fill = PatternFill(
            "solid",
            fgColor=fill_color,
        )
        value_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        value_cell.border = box_border

    # Fill the spaces between summary boxes
    for cell in ["B8", "B9", "D8", "D9", "F8", "F9"]:
        ws[cell].fill = PatternFill(
            "solid",
            fgColor=WHITE,
        )

    ws.row_dimensions[8].height = 20
    ws.row_dimensions[9].height = 28

    # ========================================================
    # CATEGORY BREAKDOWN
    # ========================================================

    ws.merge_cells("A11:G11")

    ws["A11"] = "EXPENSE CATEGORY BREAKDOWN"
    ws["A11"].font = Font(
        name="Arial",
        size=13,
        bold=True,
        color=WHITE,
    )
    ws["A11"].fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )

    category_headers = [
        "Category",
        "Transactions",
        "Total Amount",
    ]

    for col, header in enumerate(
        category_headers,
        start=1,
    ):
        cell = ws.cell(
            row=12,
            column=col,
            value=header,
        )

        cell.font = Font(
            name="Arial",
            size=10,
            bold=True,
            color=WHITE,
        )

        cell.fill = PatternFill(
            "solid",
            fgColor=NAVY_2,
        )

        cell.border = box_border

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    category_row = 13

    for item in report["expense_categories"]:

        ws.cell(
            row=category_row,
            column=1,
            value=item["category"],
        )

        ws.cell(
            row=category_row,
            column=2,
            value=item["transaction_count"],
        )

        amount_cell = ws.cell(
            row=category_row,
            column=3,
            value=item["total_amount"],
        )

        amount_cell.number_format = '₹#,##0.00'

        for col in range(1, 4):
            cell = ws.cell(
                row=category_row,
                column=col,
            )

            cell.border = box_border
            cell.alignment = Alignment(
                vertical="center",
                horizontal=(
                    "right"
                    if col in [2, 3]
                    else "left"
                ),
            )

        category_row += 1

    if not report["expense_categories"]:

        ws.cell(
            row=category_row,
            column=1,
            value="No expenses recorded",
        )

        ws.merge_cells(
            start_row=category_row,
            start_column=1,
            end_row=category_row,
            end_column=3,
        )

        ws.cell(
            row=category_row,
            column=1,
        ).font = Font(
            italic=True,
            color=SLATE,
        )

        category_row += 1

    # ========================================================
    # TRANSACTIONS
    # ========================================================

    transaction_start = category_row + 2

    ws.merge_cells(
        start_row=transaction_start,
        start_column=1,
        end_row=transaction_start,
        end_column=7,
    )

    ws.cell(
        row=transaction_start,
        column=1,
        value="TRANSACTION DETAILS",
    )

    ws.cell(
        row=transaction_start,
        column=1,
    ).font = Font(
        name="Arial",
        size=13,
        bold=True,
        color=WHITE,
    )

    ws.cell(
        row=transaction_start,
        column=1,
    ).fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )

    transaction_header_row = transaction_start + 1

    transaction_headers = [
        "Date",
        "Type",
        "Category / Source",
        "Description",
        "Payment Method",
        "Amount",
        "Account ID",
    ]

    for col, header in enumerate(
        transaction_headers,
        start=1,
    ):
        cell = ws.cell(
            row=transaction_header_row,
            column=col,
            value=header,
        )

        cell.font = Font(
            name="Arial",
            size=10,
            bold=True,
            color=WHITE,
        )

        cell.fill = PatternFill(
            "solid",
            fgColor=NAVY_2,
        )

        cell.border = box_border

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    transaction_row = transaction_header_row + 1

    for transaction in report["transactions"]:

        transaction_type = (
            transaction["type"].title()
        )

        category = (
            transaction.get("category")
            or transaction.get("source")
            or "—"
        )

        description = (
            transaction.get("description")
            or "—"
        )

        payment_method = (
            transaction.get("payment_method")
            or "—"
        )

        amount = float(
            transaction.get("amount") or 0
        )

        if transaction["type"] == "expense":
            amount = -amount

        values = [
            date_text(transaction["date"]),
            transaction_type,
            category,
            description,
            payment_method,
            amount,
            transaction.get("account_id") or "—",
        ]

        for col, value in enumerate(
            values,
            start=1,
        ):
            cell = ws.cell(
                row=transaction_row,
                column=col,
                value=value,
            )

            cell.border = box_border

            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True,
                horizontal=(
                    "right"
                    if col == 6
                    else "left"
                ),
            )

            if col == 6:
                cell.number_format = (
                    '+₹#,##0.00;-₹#,##0.00'
                )

                cell.font = Font(
                    name="Arial",
                    bold=True,
                    color=(
                        GREEN
                        if amount >= 0
                        else RED
                    ),
                )

            if col == 2:
                cell.font = Font(
                    name="Arial",
                    bold=True,
                    color=(
                        GREEN
                        if transaction_type == "Income"
                        else RED
                    ),
                )

        transaction_row += 1

    if not report["transactions"]:

        ws.merge_cells(
            start_row=transaction_row,
            start_column=1,
            end_row=transaction_row,
            end_column=7,
        )

        cell = ws.cell(
            row=transaction_row,
            column=1,
            value="No transactions recorded",
        )

        cell.font = Font(
            italic=True,
            color=SLATE,
        )

        cell.alignment = Alignment(
            horizontal="center",
        )

        transaction_row += 1

    # ========================================================
    # VERIFICATION
    # ========================================================

    verification_start = transaction_row + 2

    ws.merge_cells(
        start_row=verification_start,
        start_column=1,
        end_row=verification_start,
        end_column=7,
    )

    ws.cell(
        row=verification_start,
        column=1,
        value="REPORT VERIFICATION",
    )

    ws.cell(
        row=verification_start,
        column=1,
    ).font = Font(
        name="Arial",
        size=13,
        bold=True,
        color=WHITE,
    )

    ws.cell(
        row=verification_start,
        column=1,
    ).fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )

    verification_headers = [
        "Income Transactions",
        "Expense Transactions",
        "Total Transactions",
    ]

    verification_values = [
        report["verification"]["income_count"],
        report["verification"]["expense_count"],
        report["verification"]["transaction_count"],
    ]

    for index, (
        header,
        value,
    ) in enumerate(
        zip(
            verification_headers,
            verification_values,
        )
    ):

        start_col = 1 + index * 2

        ws.merge_cells(
            start_row=verification_start + 1,
            start_column=start_col,
            end_row=verification_start + 1,
            end_column=start_col + 1,
        )

        ws.cell(
            row=verification_start + 1,
            column=start_col,
            value=header,
        )

        ws.cell(
            row=verification_start + 1,
            column=start_col,
        ).font = Font(
            name="Arial",
            size=9,
            bold=True,
            color=NAVY,
        )

        ws.cell(
            row=verification_start + 1,
            column=start_col,
        ).fill = PatternFill(
            "solid",
            fgColor=GREEN_LIGHT,
        )

        ws.cell(
            row=verification_start + 1,
            column=start_col,
        ).alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        ws.merge_cells(
            start_row=verification_start + 2,
            start_column=start_col,
            end_row=verification_start + 2,
            end_column=start_col + 1,
        )

        ws.cell(
            row=verification_start + 2,
            column=start_col,
            value=value,
        )

        ws.cell(
            row=verification_start + 2,
            column=start_col,
        ).font = Font(
            name="Arial",
            size=14,
            bold=True,
            color=GREEN,
        )

        ws.cell(
            row=verification_start + 2,
            column=start_col,
        ).alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        for row in [
            verification_start + 1,
            verification_start + 2,
        ]:
            for col in [
                start_col,
                start_col + 1,
            ]:
                ws.cell(
                    row=row,
                    column=col,
                ).border = box_border

    # ========================================================
    # GENERAL FORMATTING
    # ========================================================

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal
                    or "left",
                    vertical="center",
                    wrap_text=True,
                )

    # Footer
    footer_row = verification_start + 5

    ws.merge_cells(
        start_row=footer_row,
        start_column=1,
        end_row=footer_row,
        end_column=7,
    )

    ws.cell(
        row=footer_row,
        column=1,
        value="Budget Buddy • Financial Report",
    )

    ws.cell(
        row=footer_row,
        column=1,
    ).font = Font(
        name="Arial",
        size=9,
        italic=True,
        color=SLATE,
    )

    ws.cell(
        row=footer_row,
        column=1,
    ).alignment = Alignment(
        horizontal="center",
    )

    # ========================================================
    # PRINT AREA
    # ========================================================

    ws.print_area = (
        f"A1:G{footer_row}"
    )

    # ========================================================
    # SAVE
    # ========================================================

    buffer = BytesIO()

    wb.save(buffer)

    buffer.seek(0)

    return buffer


