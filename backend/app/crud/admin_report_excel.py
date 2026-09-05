from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.page import PageMargins
from sqlalchemy import func
from datetime import datetime


NAVY = "071A2B"
NAVY_2 = "0B263D"
GREEN = "059669"
GREEN_LIGHT = "ECFDF5"
RED = "DC2626"
RED_LIGHT = "FEF2F2"
BLUE = "2563EB"
BLUE_LIGHT = "EFF6FF"
SLATE = "64748B"
LIGHT_SLATE = "F8FAFC"
BORDER = "CBD5E1"
WHITE = "FFFFFF"


def generate_admin_report_excel(db):
    from app.models.budget import Budget
    from app.models.expense import Expense
    from app.models.income import Income
    from app.models.savings_goal import SavingsGoal
    from app.models.user import User

    wb = Workbook()
    ws = wb.active
    ws.title = "System Report"

    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.3,
        right=0.3,
        top=0.5,
        bottom=0.5,
        header=0.2,
        footer=0.2,
    )

    widths = {
        "A": 24,
        "B": 18,
        "C": 23,
        "D": 42,
        "E": 20,
        "F": 20,
        "G": 25,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    thin = Side(style="thin", color=BORDER)

    box_border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin,
    )

    # ------------------------------------------
    # SYSTEM TOTALS
    # ------------------------------------------

    total_users = db.query(func.count(User.id)).scalar() or 0

    normal_users = (
        db.query(func.count(User.id))
        .filter(User.role == "normal")
        .scalar()
        or 0
    )

    premium_users = (
        db.query(func.count(User.id))
        .filter(User.role == "premium")
        .scalar()
        or 0
    )

    admin_users = (
        db.query(func.count(User.id))
        .filter(User.role == "admin")
        .scalar()
        or 0
    )

    active_users = (
        db.query(func.count(User.id))
        .filter(User.is_active.is_(True))
        .scalar()
        or 0
    )

    inactive_users = max(total_users - active_users, 0)

    verified_users = (
        db.query(func.count(User.id))
        .filter(User.is_verified.is_(True))
        .scalar()
        or 0
    )

    unverified_users = max(total_users - verified_users, 0)

    total_expenses = (
        db.query(func.coalesce(func.sum(Expense.amount), 0))
        .scalar()
        or 0
    )

    total_income = (
        db.query(func.coalesce(func.sum(Income.amount), 0))
        .scalar()
        or 0
    )

    total_budgets = (
        db.query(func.count(Budget.id))
        .scalar()
        or 0
    )

    total_savings_goals = (
        db.query(func.count(SavingsGoal.id))
        .scalar()
        or 0
    )

    savings_target = (
        db.query(func.coalesce(func.sum(SavingsGoal.target_amount), 0))
        .scalar()
        or 0
    )

    savings_current = (
        db.query(func.coalesce(func.sum(SavingsGoal.current_amount), 0))
        .scalar()
        or 0
    )

    net_balance = float(total_income) - float(total_expenses)

    # ------------------------------------------
    # HEADER
    # ------------------------------------------

    for row in ws["A1:G2"]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=NAVY)

    ws.merge_cells("A1:E2")
    ws["A1"] = "BUDGET BUDDY"
    ws["A1"].font = Font(
        name="Arial",
        size=24,
        bold=True,
        color=WHITE,
    )
    ws["A1"].alignment = Alignment(
        horizontal="left",
        vertical="center",
    )

    ws.merge_cells("F1:G1")
    ws["F1"] = "REPORT TYPE"
    ws["F1"].font = Font(
        name="Arial",
        size=9,
        bold=True,
        color=WHITE,
    )
    ws["F1"].fill = PatternFill("solid", fgColor=NAVY_2)
    ws["F1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws.merge_cells("F2:G2")
    ws["F2"] = "OVERALL SYSTEM REPORT"
    ws["F2"].font = Font(
        name="Arial",
        size=12,
        bold=True,
        color=NAVY,
    )
    ws["F2"].fill = PatternFill("solid", fgColor=GREEN_LIGHT)
    ws["F2"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 25

    # ------------------------------------------
    # REPORT TITLE
    # ------------------------------------------

    ws.merge_cells("A4:G4")
    ws["A4"] = "Overall System Report"
    ws["A4"].font = Font(
        name="Arial",
        size=19,
        bold=True,
        color=NAVY,
    )

    ws.merge_cells("A5:G5")
    ws["A5"] = "System-wide users, financial activity, budgets and savings overview"
    ws["A5"].font = Font(
        name="Arial",
        size=10,
        color=SLATE,
    )

    # ------------------------------------------
    # USER OVERVIEW
    # ------------------------------------------

    ws.merge_cells("A7:G7")
    ws["A7"] = "USER OVERVIEW"
    ws["A7"].font = Font(
        name="Arial",
        size=13,
        bold=True,
        color=WHITE,
    )
    ws["A7"].fill = PatternFill("solid", fgColor=NAVY)

    user_cards = [
        ("A", "B", "TOTAL USERS", total_users),
        ("C", "D", "NORMAL USERS", normal_users),
        ("E", "F", "PREMIUM USERS", premium_users),
        ("G", "G", "ADMIN USERS", admin_users),
        ("A", "B", "ACTIVE USERS", active_users),
        ("C", "D", "INACTIVE USERS", inactive_users),
        ("E", "F", "VERIFIED USERS", verified_users),
        ("G", "G", "UNVERIFIED USERS", unverified_users),
    ]

    for index, (start, end, label, value) in enumerate(user_cards):
        row = 8 if index < 4 else 10

        if start != end:
            ws.merge_cells(f"{start}{row}:{end}{row}")
            ws.merge_cells(f"{start}{row + 1}:{end}{row + 1}")

        label_cell = ws[f"{start}{row}"]
        value_cell = ws[f"{start}{row + 1}"]

        label_cell.value = label
        value_cell.value = int(value)

        label_cell.font = Font(
            name="Arial",
            size=9,
            bold=True,
            color=SLATE,
        )
        value_cell.font = Font(
            name="Arial",
            size=15,
            bold=True,
            color=NAVY,
        )

        label_cell.fill = PatternFill(
            "solid",
            fgColor=LIGHT_SLATE,
        )
        value_cell.fill = PatternFill(
            "solid",
            fgColor=WHITE,
        )

        label_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        value_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        label_cell.border = box_border
        value_cell.border = box_border

    ws.row_dimensions[8].height = 20
    ws.row_dimensions[9].height = 28
    ws.row_dimensions[10].height = 20
    ws.row_dimensions[11].height = 28

    # ------------------------------------------
    # FINANCIAL OVERVIEW
    # ------------------------------------------

    ws.merge_cells("A13:G13")
    ws["A13"] = "FINANCIAL OVERVIEW"
    ws["A13"].font = Font(
        name="Arial",
        size=13,
        bold=True,
        color=WHITE,
    )
    ws["A13"].fill = PatternFill("solid", fgColor=NAVY)

    financial_cards = [
        ("A", "B", "OVERALL INCOME", total_income, GREEN_LIGHT, GREEN),
        ("C", "D", "OVERALL EXPENSES", total_expenses, RED_LIGHT, RED),
        ("E", "F", "NET BALANCE", net_balance, BLUE_LIGHT, BLUE),
        ("G", "G", "SAVINGS CONTRIBUTIONS", savings_current, GREEN_LIGHT, GREEN),
    ]

    for start, end, label, value, fill_color, font_color in financial_cards:
        if start != end:
            ws.merge_cells(f"{start}14:{end}14")
            ws.merge_cells(f"{start}15:{end}15")

        label_cell = ws[f"{start}14"]
        value_cell = ws[f"{start}15"]

        label_cell.value = label
        value_cell.value = float(value)

        label_cell.font = Font(
            name="Arial",
            size=9,
            bold=True,
            color=font_color,
        )
        value_cell.font = Font(
            name="Arial",
            size=15,
            bold=True,
            color=font_color,
        )

        label_cell.fill = PatternFill(
            "solid",
            fgColor=fill_color,
        )
        value_cell.fill = PatternFill(
            "solid",
            fgColor=fill_color,
        )

        label_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        value_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        label_cell.border = box_border
        value_cell.border = box_border
        value_cell.number_format = '₹#,##0.00'

    ws.row_dimensions[14].height = 20
    ws.row_dimensions[15].height = 28

    # ------------------------------------------
    # PLATFORM TOTALS
    # ------------------------------------------

    ws.merge_cells("A17:G17")
    ws["A17"] = "PLATFORM TOTALS"
    ws["A17"].font = Font(
        name="Arial",
        size=13,
        bold=True,
        color=WHITE,
    )
    ws["A17"].fill = PatternFill("solid", fgColor=NAVY)

    platform_cards = [
        ("A", "B", "TOTAL BUDGETS", total_budgets, False),
        ("C", "D", "SAVINGS GOALS", total_savings_goals, False),
        ("E", "F", "SAVINGS TARGET", savings_target, True),
        ("G", "G", "CURRENT SAVINGS", savings_current, True),
    ]

    for start, end, label, value, is_currency in platform_cards:
        if start != end:
            ws.merge_cells(f"{start}18:{end}18")
            ws.merge_cells(f"{start}19:{end}19")

        label_cell = ws[f"{start}18"]
        value_cell = ws[f"{start}19"]

        label_cell.value = label
        value_cell.value = float(value) if is_currency else int(value)

        label_cell.font = Font(
            name="Arial",
            size=9,
            bold=True,
            color=SLATE,
        )
        value_cell.font = Font(
            name="Arial",
            size=15,
            bold=True,
            color=NAVY,
        )

        label_cell.fill = PatternFill(
            "solid",
            fgColor=LIGHT_SLATE,
        )
        value_cell.fill = PatternFill(
            "solid",
            fgColor=WHITE,
        )

        label_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        value_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        label_cell.border = box_border
        value_cell.border = box_border

        if is_currency:
            value_cell.number_format = '₹#,##0.00'

    ws.row_dimensions[18].height = 20
    ws.row_dimensions[19].height = 28

    # ------------------------------------------
    # MONTHLY FINANCIAL ACTIVITY
    # ------------------------------------------

    now = datetime.utcnow()
    month_starts = []

    for offset in range(5, -1, -1):
        current_month = now.month - offset
        current_year = now.year

        while current_month <= 0:
            current_month += 12
            current_year -= 1

        month_starts.append(
            datetime(current_year, current_month, 1)
        )

    ws.merge_cells("A21:G21")
    ws["A21"] = "MONTHLY FINANCIAL ACTIVITY"
    ws["A21"].font = Font(
        name="Arial",
        size=13,
        bold=True,
        color=WHITE,
    )
    ws["A21"].fill = PatternFill("solid", fgColor=NAVY)

    for column, header in enumerate(
        ["Month", "Year", "Income", "Expenses", "Net Balance"],
        start=1,
    ):
        cell = ws.cell(row=22, column=column, value=header)
        cell.font = Font(
            name="Arial",
            size=10,
            bold=True,
            color=WHITE,
        )
        cell.fill = PatternFill(
            "solid",
            fgColor=NAVY_2,
        )
        cell.border = box_border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    row = 23

    for index, start in enumerate(month_starts):
        if index < len(month_starts) - 1:
            end = month_starts[index + 1]
        elif start.month == 12:
            end = datetime(start.year + 1, 1, 1)
        else:
            end = datetime(start.year, start.month + 1, 1)

        income_total = (
            db.query(func.coalesce(func.sum(Income.amount), 0))
            .filter(
                Income.date >= start,
                Income.date < end,
            )
            .scalar()
            or 0
        )

        expense_total = (
            db.query(func.coalesce(func.sum(Expense.amount), 0))
            .filter(
                Expense.date >= start,
                Expense.date < end,
            )
            .scalar()
            or 0
        )

        values = [
            start.strftime("%b"),
            start.year,
            float(income_total),
            float(expense_total),
            float(income_total) - float(expense_total),
        ]

        for column, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=column, value=value)
            cell.border = box_border
            cell.alignment = Alignment(
                horizontal="right" if column >= 3 else "left",
                vertical="center",
            )

            if column >= 3:
                cell.number_format = '₹#,##0.00'

        row += 1

    # ------------------------------------------
    # USER REGISTRATION ACTIVITY
    # ------------------------------------------

    registration_start = row + 2

    ws.merge_cells(
        start_row=registration_start,
        start_column=1,
        end_row=registration_start,
        end_column=7,
    )

    cell = ws.cell(
        row=registration_start,
        column=1,
        value="USER REGISTRATION ACTIVITY",
    )
    cell.font = Font(
        name="Arial",
        size=13,
        bold=True,
        color=WHITE,
    )
    cell.fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )

    for column, header in enumerate(
        ["Month", "Year", "Registered Users"],
        start=1,
    ):
        cell = ws.cell(
            row=registration_start + 1,
            column=column,
            value=header,
        )
        cell.font = Font(
            name="Arial",
            size=10,
            bold=True,
            color=WHITE,
        )
        cell.fill = PatternFill(
            "solid",
            fgColor=NAVY_2,
        )
        cell.border = box_border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    row = registration_start + 2

    for index, start in enumerate(month_starts):
        if index < len(month_starts) - 1:
            end = month_starts[index + 1]
        elif start.month == 12:
            end = datetime(start.year + 1, 1, 1)
        else:
            end = datetime(start.year, start.month + 1, 1)

        registered = (
            db.query(func.count(User.id))
            .filter(
                User.created_at >= start,
                User.created_at < end,
            )
            .scalar()
            or 0
        )

        values = [
            start.strftime("%b"),
            start.year,
            registered,
        ]

        for column, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=column, value=value)
            cell.border = box_border
            cell.alignment = Alignment(
                horizontal="right" if column == 3 else "left",
                vertical="center",
            )

        row += 1

    # ------------------------------------------
    # EXPENSE CATEGORIES
    # ------------------------------------------

    expense_rows = (
        db.query(
            Expense.category,
            func.count(Expense.id),
            func.coalesce(func.sum(Expense.amount), 0),
        )
        .group_by(Expense.category)
        .order_by(func.sum(Expense.amount).desc())
        .all()
    )

    category_start = row + 2

    ws.merge_cells(
        start_row=category_start,
        start_column=1,
        end_row=category_start,
        end_column=7,
    )

    cell = ws.cell(
        row=category_start,
        column=1,
        value="EXPENSE CATEGORY BREAKDOWN",
    )
    cell.font = Font(
        name="Arial",
        size=13,
        bold=True,
        color=WHITE,
    )
    cell.fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )

    for column, header in enumerate(
        ["Category", "Transactions", "Total Amount"],
        start=1,
    ):
        cell = ws.cell(
            row=category_start + 1,
            column=column,
            value=header,
        )
        cell.font = Font(
            name="Arial",
        )
        cell.font = Font(
            name="Arial",
            size=10,
            bold=True,
            color=WHITE,
        )
        cell.fill = PatternFill(
            "solid",
            fgColor=NAVY_2,
        )
        cell.border = box_border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    row = category_start + 2

    if expense_rows:
        for category, count, amount in expense_rows:
            values = [
                category or "Other",
                count or 0,
                float(amount or 0),
            ]

            for column, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=column, value=value)
                cell.border = box_border
                cell.alignment = Alignment(
                    horizontal="right" if column > 1 else "left",
                    vertical="center",
                )

                if column == 3:
                    cell.number_format = '₹#,##0.00'

            row += 1
    else:
        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=3,
        )
        cell = ws.cell(
            row=row,
            column=1,
            value="No expenses recorded",
        )
        cell.font = Font(
            name="Arial",
            size=10,
            italic=True,
            color=SLATE,
        )
        row += 1

    # ------------------------------------------
    # INCOME CATEGORIES
    # ------------------------------------------

    income_rows = (
        db.query(
            Income.source,
            func.count(Income.id),
            func.coalesce(func.sum(Income.amount), 0),
        )
        .group_by(Income.source)
        .order_by(func.sum(Income.amount).desc())
        .all()
    )

    income_start = row + 2

    ws.merge_cells(
        start_row=income_start,
        start_column=1,
        end_row=income_start,
        end_column=7,
    )

    cell = ws.cell(
        row=income_start,
        column=1,
        value="INCOME CATEGORY BREAKDOWN",
    )
    cell.font = Font(
        name="Arial",
        size=13,
        bold=True,
        color=WHITE,
    )
    cell.fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )

    for column, header in enumerate(
        ["Source", "Transactions", "Total Amount"],
        start=1,
    ):
        cell = ws.cell(
            row=income_start + 1,
            column=column,
            value=header,
        )
        cell.font = Font(
            name="Arial",
            size=10,
            bold=True,
            color=WHITE,
        )
        cell.fill = PatternFill(
            "solid",
            fgColor=NAVY_2,
        )
        cell.border = box_border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    row = income_start + 2

    if income_rows:
        for source, count, amount in income_rows:
            values = [
                source or "Other",
                count or 0,
                float(amount or 0),
            ]

            for column, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=column, value=value)
                cell.border = box_border
                cell.alignment = Alignment(
                    horizontal="right" if column > 1 else "left",
                    vertical="center",
                )

                if column == 3:
                    cell.number_format = '₹#,##0.00'

            row += 1
    else:
        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=3,
        )
        cell = ws.cell(
            row=row,
            column=1,
            value="No income recorded",
        )
        cell.font = Font(
            name="Arial",
            size=10,
            italic=True,
            color=SLATE,
        )
        row += 1

    # ------------------------------------------
    # FOOTER
    # ------------------------------------------

    footer_row = row + 2

    ws.merge_cells(
        start_row=footer_row,
        start_column=1,
        end_row=footer_row,
        end_column=7,
    )

    cell = ws.cell(
        row=footer_row,
        column=1,
        value="Budget Buddy - Overall System Report",
    )
    cell.font = Font(
        name="Arial",
        size=10,
        bold=True,
        color=SLATE,
    )
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws.merge_cells(
        start_row=footer_row + 1,
        start_column=1,
        end_row=footer_row + 1,
        end_column=7,
    )

    cell = ws.cell(
        row=footer_row + 1,
        column=1,
        value="Admin system-wide analytics and reporting.",
    )
    cell.font = Font(
        name="Arial",
        size=9,
        color=SLATE,
    )
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws.freeze_panes = "A8"
    ws.print_title_rows = "1:7"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer