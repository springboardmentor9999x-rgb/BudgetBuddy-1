from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.crud.reports import get_report


def generate_report_excel_limited(
    db,
    user_id: int,
    period: str,
    start_date=None,
    end_date=None,
):
    report = get_report(
        db=db,
        user_id=user_id,
        period=period,
        start_date=start_date,
        end_date=end_date,
    )

    transactions = report.get("transactions", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Title
    ws["A1"] = "Budget Buddy"
    ws["A1"].font = Font(size=18, bold=True)

    ws["A2"] = "Transaction Report"
    ws["A2"].font = Font(size=14, bold=True)

    ws["A3"] = (
        f"Period: {report.get('start_date')} - "
        f"{report.get('end_date')}"
    )

    # Headers
    headers = [
        "Date",
        "Type",
        "Category / Source",
        "Description",
        "Payment Method",
        "Amount",
    ]

    header_row = 5

    for column, header in enumerate(headers, start=1):
        cell = ws.cell(
            row=header_row,
            column=column,
            value=header,
        )
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            "solid",
            fgColor="0F172A",
        )
        cell.alignment = Alignment(
            horizontal="center"
        )

    # Transactions
    row = header_row + 1

    for transaction in transactions:
        transaction_type = (
            str(transaction.get("type", "")).title()
        )

        category = (
            transaction.get("category")
            or transaction.get("source")
            or "-"
        )

        description = (
            transaction.get("description")
            or "-"
        )

        payment_method = (
            transaction.get("payment_method")
            or "-"
        )

        amount = (
            transaction.get("amount")
            or 0
        )

        values = [
            transaction.get("date"),
            transaction_type,
            category,
            description,
            payment_method,
            amount,
        ]

        for column, value in enumerate(
            values,
            start=1,
        ):
            ws.cell(
                row=row,
                column=column,
                value=value,
            )

        row += 1

    if not transactions:
        ws.cell(
            row=row,
            column=1,
            value="No transactions recorded.",
        )

    # Formatting
    widths = {
        "A": 18,
        "B": 14,
        "C": 24,
        "D": 35,
        "E": 20,
        "F": 16,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    thin = Side(
        style="thin",
        color="D1D5DB",
    )

    for current_row in ws.iter_rows(
        min_row=header_row,
        max_row=max(row - 1, header_row),
        min_col=1,
        max_col=len(headers),
    ):
        for cell in current_row:
            cell.border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin,
            )

    ws.freeze_panes = "A6"

    # Save
    buffer = BytesIO()

    wb.save(buffer)
    buffer.seek(0)

    return buffer
